"""
指纹解锁速度分析引擎 - 从原 CLI 脚本抽取的复用模块
"""

import os
import re
import glob
import zipfile
import subprocess
import traceback
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    Workbook = None

# ===== 配置常量 =====
EXTRACTED_DIR = "logs_extracted"
FILTERED_DIR = "filtered_logs"
OUTPUT_DIR = "analysis_results"
PASSWORD = "ADe88sWMJt8P4QCA2E^VNacbFtY6cOdB"

# ===== 过滤关键字 =====
FILTER_KEYWORDS = [
    "gf_hal",
    "noth-aidl",
    "EVENT_FINGER_DOWN",
    "Setting power mode",
    "Finished setting power mode",
    "EVENT_UI_READY",
    "Begin capture after",
    "Auth success",
    "AuthenticationClient: onAuthenticated",
    "SURFACE SHOW",
    "Time from finger down to success notification",
    "KeyguardViewMediator",
    "KPI time",
]

FILTER_PATTERN = re.compile("|".join(re.escape(kw) for kw in FILTER_KEYWORDS), re.IGNORECASE)

# ===== 时间戳 & 数值提取 =====
TS_PATTERN = re.compile(r"^(\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}\.\d{3})")
MS_PATTERN = re.compile(r"=\s*(\d+)\s*ms")
CAPTURE_MS_PATTERN = re.compile(r"Begin capture after.*?(\d+)\s*ms")
POWER_MODE_RE = re.compile(r"Setting power mode (\d+)")

# ===== 解锁步骤定义 =====
STEPS = [
    ("EVENT_FINGER_DOWN",       "手指按压",               "timestamp"),
    ("Setting power mode 2",    "屏幕切状态(亮屏)",        "timestamp"),
    ("Finished setting power mode 2", "屏幕完成切状态",    "timestamp"),
    ("EVENT_UI_READY",          "光斑高亮",               "timestamp"),
    ("Begin capture after",     "指纹开始抓图(ms)",        "extract_ms_capture"),
    ("Auth success",            "指纹匹配成功",            "timestamp"),
    ("onAuthenticated",         "UI收到匹配成功",          "timestamp"),
    ("Time from finger down to success notification", "HAL按压到解锁(ms)", "extract_ms"),
    ("KPI time",                "HAL算法时间(ms)",         "extract_ms"),
    ("KeyguardViewMediator: exitKeyguardAndFinishSurfaceBehindRemoteAnimation", "开始显示应用画面", "timestamp"),
    ("EVENT_FINGER_UP",         "手指抬起",               "timestamp"),
]


# ===== 工具函数 =====

def detect_encoding(filepath: str) -> str:
    """检测文件编码 (BOM 检测)"""
    with open(filepath, "rb") as f:
        head = f.read(4)
    if head[:2] == b"\xff\xfe":
        return "utf-16-le"
    elif head[:2] == b"\xfe\xff":
        return "utf-16-be"
    elif head[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    return "utf-8"


def parse_timestamp(line: str) -> datetime | None:
    """从日志行提取时间戳"""
    m = TS_PATTERN.match(line)
    if m:
        try:
            return datetime.strptime(f"2026-{m.group(1)}", "%Y-%m-%d %H:%M:%S.%f")
        except ValueError:
            return None
    return None


def extract_ms_value(line: str) -> str | None:
    """提取 = xxx ms 格式的数值"""
    m = MS_PATTERN.search(line)
    return m.group(1) if m else None


def calc_delta_ms(t1, t2) -> str:
    """计算两个 datetime 的毫秒差"""
    if not isinstance(t1, datetime) or not isinstance(t2, datetime):
        return "-"
    return f"{(t2 - t1).total_seconds() * 1000:.1f}"


def format_ts(val) -> str:
    """格式化时间戳显示"""
    if val is None:
        return "-"
    if isinstance(val, datetime):
        return val.strftime("%H:%M:%S.%f")[:-3]
    return str(val)


def _safe_float(val) -> float | None:
    """安全转浮点"""
    if val is None or val == "-":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ===== 核心分析函数 =====

def analyze_log_file(filepath: str) -> list[dict]:
    """分析单个过滤后的日志文件，返回解锁记录列表"""
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    unlocks = []
    current = None
    last_power_mode = None

    for line in lines:
        pm = POWER_MODE_RE.search(line)
        if pm and "Finished" not in line:
            last_power_mode = int(pm.group(1))

        if "EVENT_FINGER_DOWN" in line:
            if current is not None:
                unlocks.append(current)
            current = {step[1]: None for step in STEPS}
            current["手指按压"] = parse_timestamp(line)
            current["_pre_power_mode"] = last_power_mode
            current["_has_pm2_between"] = False
            continue

        if current is None:
            continue

        if pm and "Finished" not in line and int(pm.group(1)) == 2:
            if current.get("UI收到匹配成功") is None:
                current["_has_pm2_between"] = True

        for keyword, label, mode in STEPS[1:]:
            if current[label] is not None:
                continue
            if keyword not in line:
                continue
            if mode == "timestamp":
                current[label] = parse_timestamp(line)
            elif mode == "extract_ms":
                current[label] = extract_ms_value(line)
            elif mode == "extract_ms_capture":
                m = CAPTURE_MS_PATTERN.search(line)
                current[label] = m.group(1) if m else None
            break

    if current is not None:
        unlocks.append(current)

    # 分类解锁类型
    for u in unlocks:
        has_pm2 = u.get("_has_pm2_between", False)
        pre_pm = u.get("_pre_power_mode")
        if has_pm2 and pre_pm == 0:
            u["_unlock_type"] = "息屏解锁"
        elif has_pm2 and pre_pm == 1:
            u["_unlock_type"] = "AOD解锁"
        elif not has_pm2 and pre_pm == 2:
            u["_unlock_type"] = "亮屏解锁"
        else:
            u["_unlock_type"] = "亮屏解锁"

    return unlocks


def filter_log_file(input_path: str, output_path: str) -> tuple[int, int]:
    """过滤单个日志文件，返回 (总行数, 匹配行数)"""
    enc = detect_encoding(input_path)
    matched = 0
    total = 0
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(input_path, "r", encoding=enc, errors="ignore") as fin, \
         open(output_path, "w", encoding="utf-8") as fout:
        for line in fin:
            total += 1
            if FILTER_PATTERN.search(line):
                fout.write(line)
                matched += 1
    return total, matched


def find_android_logs(base_dir: str) -> list[str]:
    """在解压目录中递归查找 general_log/android.log"""
    results = []
    for root, dirs, files in os.walk(base_dir):
        if "general_log" in root and "android.log" in files:
            results.append(os.path.join(root, "android.log"))
    results.sort()
    return results


def extract_zip_files(zip_files: list[str]) -> list[str]:
    """解压 zip 文件，返回所有找到的 android.log 路径列表"""

    os.makedirs(EXTRACTED_DIR, exist_ok=True)

    sz_paths = [
        "7z",
        r"C:\Program Files\7-Zip\7z.exe",
        r"D:\Program Files\7-Zip\7z.exe",
    ]
    sz_exe = None
    for p in sz_paths:
        try:
            subprocess.run([p, "--help"], capture_output=True, timeout=5)
            sz_exe = p
            break
        except (FileNotFoundError, subprocess.TimeoutExpired):
            continue

    all_logs = []
    for zf_path in zip_files:
        zf_name = os.path.basename(zf_path)
        dest_dir = os.path.join(EXTRACTED_DIR, os.path.splitext(zf_name)[0])
        os.makedirs(dest_dir, exist_ok=True)

        if sz_exe:
            cmd = [sz_exe, "x", f"-p{PASSWORD}", f"-o{dest_dir}", "-mmt=on", "-aoa", zf_path]
            result = subprocess.run(cmd, capture_output=True, text=True)
            if result.returncode != 0:
                raise RuntimeError(f"7z 解压失败 ({zf_name}): {result.stderr.strip()}")
        else:
            try:
                with zipfile.ZipFile(zf_path, "r") as zf:
                    zf.extractall(path=dest_dir, pwd=PASSWORD.encode("utf-8"))
            except RuntimeError as e:
                raise RuntimeError(f"解压失败 (密码错误?): {zf_name} - {e}")
            except zipfile.BadZipFile:
                raise RuntimeError(f"文件损坏: {zf_name}")

        # 解压后查找 android.log
        logs = find_android_logs(dest_dir)
        if not logs:
            for root, dirs, files in os.walk(dest_dir):
                if "android.log" in files:
                    logs.append(os.path.join(root, "android.log"))
            logs.sort()
        all_logs.extend(logs)

    return all_logs


def write_excel(all_results: dict[str, list[dict]], output_path: str):
    """生成 Excel 报告"""
    if Workbook is None:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    screen_on_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    aod_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    avg_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_bold = Font(bold=True, color="FF0000")
    blue_bold = Font(bold=True, color="0070C0")
    avg_font = Font(bold=True, size=11)

    headers = ["解锁次数", "解锁类型"]
    for i, (kw, label, mode) in enumerate(STEPS):
        if mode == "timestamp":
            headers.append(f"{label}\n(时间戳)")
            if i > 0:
                headers.append("距上一步(ms)")
        elif mode in ("extract_ms", "extract_ms_capture"):
            headers.append(label)
    headers.append("底层时间(ms)\nFINGER_DOWN→Auth success")
    headers.append("总解锁时间(ms)\nFINGER_DOWN→开始显示应用画面")

    for log_name, unlocks in all_results.items():
        ws = wb.create_sheet(title=log_name[:31])

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        if not unlocks:
            for col_cells in ws.columns:
                max_len = 0
                for cell in col_cells:
                    if cell.value:
                        for line in str(cell.value).split("\n"):
                            max_len = max(max_len, len(line) + 2)
                ws.column_dimensions[col_cells[0].column_letter].width = max(max_len * 1.3, 12)
            ws.freeze_panes = "C2"
            continue

        row_data_by_type = {}

        for idx, unlock in enumerate(unlocks, 1):
            row = idx + 1
            finger_down_ts = unlock.get("手指按压")
            auth_success_ts = unlock.get("指纹匹配成功")
            app_vis_ts = unlock.get("开始显示应用画面")
            unlock_type = unlock.get("_unlock_type", "未知")

            row_vals = {}
            col = 1

            c = ws.cell(row=row, column=col, value=f"第{idx}次")
            c.border = thin_border; c.alignment = center
            col += 1
            c = ws.cell(row=row, column=col, value=unlock_type)
            c.border = thin_border; c.alignment = center
            if unlock_type == "亮屏解锁":
                c.fill = screen_on_fill
            elif unlock_type == "AOD解锁":
                c.fill = aod_fill
            col += 1

            prev_ts = None
            for i, (kw, label, mode) in enumerate(STEPS):
                val = unlock.get(label)
                if mode == "timestamp":
                    c = ws.cell(row=row, column=col, value=format_ts(val))
                    c.border = thin_border; c.alignment = center
                    col += 1
                    if i > 0:
                        delta = calc_delta_ms(prev_ts, val) if isinstance(val, datetime) else "-"
                        c = ws.cell(row=row, column=col, value=delta)
                        c.border = thin_border; c.alignment = center
                        fv = _safe_float(delta)
                        if fv is not None:
                            row_vals[col] = fv
                        col += 1
                    if isinstance(val, datetime):
                        prev_ts = val
                elif mode in ("extract_ms", "extract_ms_capture"):
                    c = ws.cell(row=row, column=col, value=val if val else "-")
                    c.border = thin_border; c.alignment = center
                    fv = _safe_float(val)
                    if fv is not None:
                        row_vals[col] = fv
                    col += 1

            base_time = calc_delta_ms(finger_down_ts, auth_success_ts)
            c = ws.cell(row=row, column=col, value=base_time)
            c.border = thin_border; c.alignment = center
            if base_time != "-":
                c.font = blue_bold
            fv = _safe_float(base_time)
            if fv is not None:
                row_vals[col] = fv
            base_col = col
            col += 1

            total_time = calc_delta_ms(finger_down_ts, app_vis_ts)
            c = ws.cell(row=row, column=col, value=total_time)
            c.border = thin_border; c.alignment = center
            if total_time != "-":
                c.font = red_bold
            elif unlock_type == "亮屏解锁":
                c.value = "亮屏解锁(无画面切换)"
                c.fill = screen_on_fill
            fv = _safe_float(total_time)
            if fv is not None:
                row_vals[col] = fv
            total_col = col

            row_data_by_type.setdefault(unlock_type, []).append(row_vals)

        max_col = total_col
        current_row = len(unlocks) + 3

        for utype in ["息屏解锁", "AOD解锁", "亮屏解锁"]:
            rows_of_type = row_data_by_type.get(utype, [])
            if not rows_of_type:
                continue

            c = ws.cell(row=current_row, column=1, value=f"{utype} 平均值 ({len(rows_of_type)}次)")
            c.font = avg_font
            c.fill = avg_fill
            c.border = thin_border
            c.alignment = center
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            c2 = ws.cell(row=current_row, column=2)
            c2.border = thin_border

            for col_idx in range(3, max_col + 1):
                vals = [r.get(col_idx) for r in rows_of_type if r.get(col_idx) is not None]
                if vals:
                    avg = sum(vals) / len(vals)
                    c = ws.cell(row=current_row, column=col_idx, value=f"{avg:.1f}")
                else:
                    c = ws.cell(row=current_row, column=col_idx, value="-")
                c.font = avg_font
                c.fill = avg_fill
                c.border = thin_border
                c.alignment = center

            current_row += 1

        for col_cells in ws.columns:
            max_len = 0
            for cell in col_cells:
                if cell.value:
                    for line in str(cell.value).split("\n"):
                        max_len = max(max_len, len(line) + 2)
            ws.column_dimensions[col_cells[0].column_letter].width = max(max_len * 1.3, 12)
        ws.freeze_panes = "C2"

    wb.save(output_path)


def run_analysis(log_files: list[str]) -> dict[str, list[dict]]:
    """运行完整分析流程（含 zip 解压 + 过滤 + 分析）"""
    analysis_logs = []

    # 分离 zip 和 log/txt 文件
    zip_files = []
    direct_logs = []
    for f in log_files:
        if not os.path.isfile(f):
            analysis_logs.append(f"WARN: 文件不存在: {f}")
            continue
        ext = os.path.splitext(f)[1].lower()
        if ext == ".zip":
            zip_files.append(f)
        else:
            direct_logs.append(f)

    all_results = {}

    # 处理 zip 文件：解压 → 查找 android.log → 过滤 → 分析
    if zip_files:
        analysis_logs.append(f"--- 解压 {len(zip_files)} 个 zip 文件 ---")
        extracted_logs = extract_zip_files(zip_files)
        analysis_logs.append(f"解压后找到 {len(extracted_logs)} 个 android.log")
        for log_path in extracted_logs:
            rel = os.path.relpath(log_path, EXTRACTED_DIR)
            parts = rel.replace("\\", "/").split("/")
            zip_name = parts[0] if parts else "unknown"
            ntlog_name = parts[-2] if len(parts) >= 2 else "unknown"
            out_filename = f"{zip_name}__{ntlog_name}__android_fingerprint_filtered.log"
            dst_path = os.path.join(FILTERED_DIR, out_filename)
            total, matched = filter_log_file(log_path, dst_path)
            analysis_logs.append(f"  {zip_name}/{ntlog_name}: {total}行 -> 匹配{matched}行")

            unlocks = analyze_log_file(dst_path)
            valid = [u for u in unlocks if u.get("指纹匹配成功") is not None]
            analysis_logs.append(f"    检测到 {len(unlocks)} 次按压, {len(valid)} 次成功解锁")
            all_results[f"{zip_name}/{ntlog_name}"] = valid

    # 处理 log/txt 文件：直接过滤 → 分析
    if direct_logs:
        analysis_logs.append(f"--- 处理 {len(direct_logs)} 个 log/txt 文件 ---")
        for lf in sorted(direct_logs):
            basename = os.path.basename(lf)
            short_name = os.path.splitext(basename)[0]
            analysis_logs.append(f"  {basename}")

            filtered_dir = os.path.join(os.path.dirname(lf) or ".", FILTERED_DIR)
            out_filename = f"{short_name}__android_fingerprint_filtered.log"
            dst_path = os.path.join(filtered_dir, out_filename)
            total, matched = filter_log_file(lf, dst_path)
            analysis_logs.append(f"    过滤: {total} 行 -> 匹配 {matched} 行")

            unlocks = analyze_log_file(dst_path)
            valid = [u for u in unlocks if u.get("指纹匹配成功") is not None]
            analysis_logs.append(f"    检测到 {len(unlocks)} 次按压, {len(valid)} 次成功解锁")
            all_results[short_name] = valid

    if all_results:
        all_results["__analysis_log__"] = analysis_logs
    return all_results