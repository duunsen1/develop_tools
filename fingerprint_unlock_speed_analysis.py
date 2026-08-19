"""
安卓指纹解锁日志一键分析工具
一条命令完成: 解压 -> 过滤关键字 -> 分析时间 -> 生成Excel

用法:
  python fingerprint_unlock_speed_analysis.py <file1.zip> [file2.zip ...]     # 解压+分析
  python fingerprint_unlock_speed_analysis.py <file1.log> [file2.txt ...]    # 直接分析
  python fingerprint_unlock_speed_analysis.py <file1.zip> <file2.log> ...    # 混合模式
"""

import zipfile
import subprocess
import os
import sys
import re
import glob
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# Windows 下隐藏子进程(7z等)的控制台窗口
CREATE_NO_WINDOW = getattr(subprocess, "CREATE_NO_WINDOW", 0)

# ========== 配置区 ==========
EXTRACTED_DIR = "logs_extracted"
FILTERED_DIR = "filtered_logs"
OUTPUT_DIR = "analysis_results"
PASSWORD = "ADe88sWMJt8P4QCA2E^VNacbFtY6cOdB"  # TODO: 在这里填入解压密码
# ============================

# ============================================================
#  第一步: 解压 logs 目录下的 zip 文件
# ============================================================

def step1_extract(zip_files: list[str]):
    print("=" * 60)
    print(" 第一步: 解压日志文件")
    print("=" * 60)

    if not PASSWORD:
        print("错误: 请先在脚本中填入解压密码 (PASSWORD 变量)")
        sys.exit(1)

    os.makedirs(EXTRACTED_DIR, exist_ok=True)

    # 查找 7z
    sz_paths = [
        "7z",
        r"C:\Program Files\7-Zip\7z.exe",
        r"D:\Program Files\7-Zip\7z.exe",
    ]
    sz_exe = None
    for p in sz_paths:
        try:
            subprocess.run([p, "--help"], capture_output=True, timeout=5, creationflags=CREATE_NO_WINDOW)
            sz_exe = p
            break
        except (FileNotFoundError, subprocess.TimeoutExpired):
            continue

    for zf_path in zip_files:
        zf_name = os.path.basename(zf_path)
        dest_dir = os.path.join(EXTRACTED_DIR, os.path.splitext(zf_name)[0])
        os.makedirs(dest_dir, exist_ok=True)

        print(f"正在解压: {zf_name}")

        if sz_exe:
            # 7z 多线程解压: x=保留目录结构, -mmt=on 多线程, -aoa 覆盖
            cmd = [sz_exe, "x", f"-p{PASSWORD}", f"-o{dest_dir}", "-mmt=on", "-aoa", zf_path]
            result = subprocess.run(cmd, capture_output=True, text=True, creationflags=CREATE_NO_WINDOW)
            if result.returncode != 0:
                print(f"  7z 解压失败: {result.stderr.strip()}")
                continue
            count = sum(len(files) for _, _, files in os.walk(dest_dir))
            print(f"  7z 解压完成, 共 {count} 个文件")
        else:
            # fallback: Python zipfile
            print("  (未找到7z, 使用Python zipfile解压)")
            try:
                with zipfile.ZipFile(zf_path, "r") as zf:
                    zf.extractall(path=dest_dir, pwd=PASSWORD.encode("utf-8"))
                count = sum(len(files) for _, _, files in os.walk(dest_dir))
                print(f"  解压完成, 共 {count} 个文件")
            except RuntimeError as e:
                print(f"  解压失败 (密码错误?): {e}")
            except zipfile.BadZipFile:
                print(f"  文件损坏，跳过: {zf_name}")

    print(f"\n解压完成 -> {EXTRACTED_DIR}/\n")


# ============================================================
#  第二步: 过滤 android.log 中的指纹相关关键字
# ============================================================

FILTER_KEYWORDS = [
    "gf_hal",                                    # 汇顶指纹HAL日志
    "noth-aidl",                                 # Nothing 指纹AIDL服务log
    "EVENT_FINGER_DOWN",                         # 手指按压
    "Setting power mode",                        # 屏幕切状态
    "Finished setting power mode",               # 屏幕完成切状态动作
    "EVENT_UI_READY",                            # 光斑高亮
    "Begin capture after",                       # 指纹开始抓图片
    "Auth success",                              # 指纹匹配成功
    "AuthenticationClient: onAuthenticated",     # UI接收到指纹匹配成功消息
    "SURFACE SHOW",                              # 进入到桌面
    "Time from finger down to success notification",  # HAL统计按压到解锁时间
    "KeyguardViewMediator",                      # 开始显示应用画面
    "WMLsVisInteractor",                         # 开始显示应用画面(新项目)
    "KPI time",                                  # 指纹HAL算法时间
]

FILTER_PATTERN = re.compile("|".join(re.escape(kw) for kw in FILTER_KEYWORDS), re.IGNORECASE)


# 匹配 android.log 及滚动日志 (android.log.N)
ANDROID_LOG_RE = re.compile(r"^android\.log(\.\d+)?$", re.IGNORECASE)


def find_android_logs(base_dir: str) -> list[str]:
    results = []
    for root, dirs, files in os.walk(base_dir):
        if "general_log" not in root:
            continue
        for f in files:
            if ANDROID_LOG_RE.match(f):
                results.append(os.path.join(root, f))
    results.sort()
    return results


def _concat_order(log_path: str):
    """滚动日志按时间排序(升序=旧到新): android.log.N 数值越大越旧, android.log 最新"""
    base = os.path.basename(log_path)
    m = re.match(r"android\.log(?:\.(\d+))?$", base, re.IGNORECASE)
    if m and m.group(1):
        return -int(m.group(1))
    return 0  # android.log 为最新，排最后


def step2_filter():
    print("=" * 60)
    print(" 第二步: 过滤指纹相关日志")
    print("=" * 60)

    os.makedirs(FILTERED_DIR, exist_ok=True)
    android_logs = find_android_logs(EXTRACTED_DIR)
    if not android_logs:
        print(f"在 {EXTRACTED_DIR} 中未找到 general_log/android.log")
        return

    print(f"找到 {len(android_logs)} 个 android.log 文件(含滚动):\n")

    # 按压缩包分组，同一压缩包的滚动日志按时间顺序合并，共用一个 sheet
    by_zip = {}
    for log_path in android_logs:
        rel = os.path.relpath(log_path, EXTRACTED_DIR)
        parts = rel.replace("\\", "/").split("/")
        zip_name = parts[0] if len(parts) > 0 else "unknown"
        by_zip.setdefault(zip_name, []).append(log_path)

    for zip_name, logs in by_zip.items():
        logs.sort(key=_concat_order)
        dst_path = os.path.join(FILTERED_DIR, f"{zip_name}__android_fingerprint_filtered.log")

        matched = 0
        total = 0
        with open(dst_path, "w", encoding="utf-8") as fout:
            for log_path in logs:
                print(f"处理: {log_path}")
                with open(log_path, "r", encoding="utf-8", errors="ignore") as fin:
                    for line in fin:
                        total += 1
                        if FILTER_PATTERN.search(line):
                            fout.write(line)
                            matched += 1
        print(f"  {zip_name}: 合并 {len(logs)} 个日志 -> {total} 行 -> 匹配 {matched} 行 -> {dst_path}")

    print(f"\n过滤完成 -> {FILTERED_DIR}/\n")


# ============================================================
#  第三步: 分析指纹解锁时间并生成 Excel
# ============================================================

TS_PATTERN = re.compile(r"^(\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}\.\d{3})")
MS_PATTERN = re.compile(r"=\s*(\d+)\s*ms")
CAPTURE_MS_PATTERN = re.compile(r"Begin capture after.*?(\d+)\s*ms")

STEPS = [
    ("EVENT_FINGER_DOWN",       "手指按压",               "timestamp"),
    ("Setting power mode 2",    "屏幕切状态(亮屏)",        "timestamp"),
    ("Finished setting power mode 2", "屏幕完成切状态",    "timestamp"),
    ("EVENT_UI_READY",          "光斑高亮",               "timestamp"),
    ("Begin capture after",     "指纹开始抓图(ms)",        "extract_ms_capture"),
    ("Auth success",            "指纹匹配成功",            "timestamp"),
    ("onAuthenticated",         "UI收到匹配成功",          "timestamp"),
    ("Time from finger down to success notification", "HAL按压到解锁(ms)", "extract_ms"),
    ("KPI time",                "HAL算法时间(ms)",         "extract_ms"),
    (("surfaceBehindVisibility=true", "exitKeyguardAndFinishSurfaceBehindRemoteAnimation"), "开始显示应用画面", "timestamp"),
    ("EVENT_FINGER_UP",         "手指抬起",               "timestamp"),
]


def parse_timestamp(line: str) -> datetime | None:
    m = TS_PATTERN.match(line)
    if m:
        ts_str = m.group(1)
        try:
            return datetime.strptime(f"2026-{ts_str}", "%Y-%m-%d %H:%M:%S.%f")
        except ValueError:
            return None
    return None


def extract_ms_value(line: str) -> str | None:
    m = MS_PATTERN.search(line)
    return m.group(1) if m else None


POWER_MODE_RE = re.compile(r"Setting power mode (\d+)")


def analyze_log_file(filepath: str) -> list[dict]:
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    unlocks = []
    current = None

    # 追踪 FINGER_DOWN 前最近一次 Setting power mode 的数字
    last_power_mode = None

    for line in lines:
        # 持续追踪 Setting power mode X
        pm = POWER_MODE_RE.search(line)
        if pm and "Finished" not in line:
            last_power_mode = int(pm.group(1))

        if "EVENT_FINGER_DOWN" in line:
            if current is not None:
                unlocks.append(current)
            current = {step[1]: None for step in STEPS}
            current["手指按压"] = parse_timestamp(line)
            current["_pre_power_mode"] = last_power_mode  # 记录 FINGER_DOWN 前的 power mode
            current["_has_pm2_between"] = False  # FINGER_DOWN 到 onAuthenticated 之间是否有 Setting power mode 2
            continue

        if current is None:
            continue

        # 追踪 FINGER_DOWN 到 onAuthenticated 之间是否出现 Setting power mode 2
        if pm and "Finished" not in line and int(pm.group(1)) == 2:
            if current.get("UI收到匹配成功") is None:  # 还没到 onAuthenticated
                current["_has_pm2_between"] = True

        for keyword, label, mode in STEPS[1:]:
            if current[label] is not None:
                continue
            keywords = keyword if isinstance(keyword, (tuple, list)) else (keyword,)
            if not any(k in line for k in keywords):
                continue
            if mode == "timestamp":
                current[label] = parse_timestamp(line)
            elif mode == "extract_ms":
                current[label] = extract_ms_value(line)
            elif mode == "extract_ms_capture":
                m = CAPTURE_MS_PATTERN.search(line)
                current[label] = m.group(1) if m else None
            break

    if current is not None:
        unlocks.append(current)

    # 判断解锁类型
    for u in unlocks:
        has_pm2 = u.get("_has_pm2_between", False)
        pre_pm = u.get("_pre_power_mode")

        if has_pm2 and pre_pm == 0:
            u["_unlock_type"] = "息屏解锁"
        elif has_pm2 and pre_pm == 1:
            u["_unlock_type"] = "AOD解锁"
        elif not has_pm2 and pre_pm == 2:
            u["_unlock_type"] = "亮屏解锁"
        else:
            u["_unlock_type"] = "亮屏解锁"

    return unlocks


def calc_delta_ms(t1, t2) -> str:
    if not isinstance(t1, datetime) or not isinstance(t2, datetime):
        return "-"
    return f"{(t2 - t1).total_seconds() * 1000:.1f}"


def format_ts(val) -> str:
    if val is None:
        return "-"
    if isinstance(val, datetime):
        return val.strftime("%H:%M:%S.%f")[:-3]
    return str(val)


def _safe_float(val) -> float | None:
    """尝试将值转为 float，失败返回 None"""
    if val is None or val == "-":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def write_excel(all_results: dict[str, list[dict]], output_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    screen_on_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    aod_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    avg_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_bold = Font(bold=True, color="FF0000")
    blue_bold = Font(bold=True, color="0070C0")
    avg_font = Font(bold=True, size=11)

    headers = ["解锁次数", "解锁类型"]
    for i, (kw, label, mode) in enumerate(STEPS):
        if mode == "timestamp":
            headers.append(f"{label}\n(时间戳)")
            if i > 0:
                headers.append("距上一步(ms)")
        elif mode in ("extract_ms", "extract_ms_capture"):
            headers.append(label)
    headers.append("底层时间(ms)\nFINGER_DOWN→Auth success")
    headers.append("总解锁时间(ms)\nFINGER_DOWN→开始显示应用画面")

    # 记录哪些列是数值列（用于计算平均值）
    # 数值列: 距上一步(ms)、extract_ms 类、底层时间、总解锁时间
    num_col_indices = []  # (col_index, label) 列表，1-based

    for log_name, unlocks in all_results.items():
        ws = wb.create_sheet(title=log_name[:31])

        # 写表头
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # 如果没有有效解锁数据，跳过数据行，只保留表头
        if not unlocks:
            for col_cells in ws.columns:
                max_len = 0
                for cell in col_cells:
                    if cell.value:
                        for line in str(cell.value).split("\n"):
                            max_len = max(max_len, len(line) + 2)
                ws.column_dimensions[col_cells[0].column_letter].width = max(max_len * 1.3, 12)
            ws.freeze_panes = "C2"
            continue

        # 收集每行的数值数据，用于后续按类型统计
        # row_data[idx] = { col_index: float_value, ... }
        row_data_by_type = {}  # { unlock_type: [ {col: val, ...}, ... ] }

        for idx, unlock in enumerate(unlocks, 1):
            row = idx + 1
            finger_down_ts = unlock.get("手指按压")
            auth_success_ts = unlock.get("指纹匹配成功")
            app_vis_ts = unlock.get("开始显示应用画面")
            unlock_type = unlock.get("_unlock_type", "未知")

            row_vals = {}  # col_index -> float

            col = 1
            c = ws.cell(row=row, column=col, value=f"第{idx}次")
            c.border = thin_border; c.alignment = center
            col += 1
            c = ws.cell(row=row, column=col, value=unlock_type)
            c.border = thin_border; c.alignment = center
            if unlock_type == "亮屏解锁":
                c.fill = screen_on_fill
            elif unlock_type == "AOD解锁":
                c.fill = aod_fill
            col += 1

            prev_ts = None
            for i, (kw, label, mode) in enumerate(STEPS):
                val = unlock.get(label)
                if mode == "timestamp":
                    c = ws.cell(row=row, column=col, value=format_ts(val))
                    c.border = thin_border; c.alignment = center
                    col += 1
                    if i > 0:
                        delta = calc_delta_ms(prev_ts, val) if isinstance(val, datetime) else "-"
                        c = ws.cell(row=row, column=col, value=delta)
                        c.border = thin_border; c.alignment = center
                        fv = _safe_float(delta)
                        if fv is not None:
                            row_vals[col] = fv
                        col += 1
                    if isinstance(val, datetime):
                        prev_ts = val
                elif mode in ("extract_ms", "extract_ms_capture"):
                    c = ws.cell(row=row, column=col, value=val if val else "-")
                    c.border = thin_border; c.alignment = center
                    fv = _safe_float(val)
                    if fv is not None:
                        row_vals[col] = fv
                    col += 1

            base_time = calc_delta_ms(finger_down_ts, auth_success_ts)
            c = ws.cell(row=row, column=col, value=base_time)
            c.border = thin_border; c.alignment = center
            if base_time != "-":
                c.font = blue_bold
            fv = _safe_float(base_time)
            if fv is not None:
                row_vals[col] = fv
            base_col = col
            col += 1

            total_time = calc_delta_ms(finger_down_ts, app_vis_ts)
            c = ws.cell(row=row, column=col, value=total_time)
            c.border = thin_border; c.alignment = center
            if total_time != "-":
                c.font = red_bold
            elif unlock_type == "亮屏解锁":
                c.value = "亮屏解锁(无画面切换)"
                c.fill = screen_on_fill
            fv = _safe_float(total_time)
            if fv is not None:
                row_vals[col] = fv
            total_col = col

            row_data_by_type.setdefault(unlock_type, []).append(row_vals)

        # ===== 分类统计平均值 =====
        max_col = total_col
        current_row = len(unlocks) + 3  # 空一行

        for utype in ["息屏解锁", "AOD解锁", "亮屏解锁"]:
            rows_of_type = row_data_by_type.get(utype, [])
            if not rows_of_type:
                continue

            # 标题行
            c = ws.cell(row=current_row, column=1, value=f"{utype} 平均值 ({len(rows_of_type)}次)")
            c.font = avg_font
            c.fill = avg_fill
            c.border = thin_border
            c.alignment = center
            # 合并前两列作为标题
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            c2 = ws.cell(row=current_row, column=2)
            c2.border = thin_border

            # 计算每列平均值
            for col_idx in range(3, max_col + 1):
                vals = [r.get(col_idx) for r in rows_of_type if r.get(col_idx) is not None]
                if vals:
                    avg = sum(vals) / len(vals)
                    c = ws.cell(row=current_row, column=col_idx, value=f"{avg:.1f}")
                else:
                    c = ws.cell(row=current_row, column=col_idx, value="-")
                c.font = avg_font
                c.fill = avg_fill
                c.border = thin_border
                c.alignment = center

            current_row += 1

        # 自动列宽
        for col_cells in ws.columns:
            max_len = 0
            for cell in col_cells:
                if cell.value:
                    for line in str(cell.value).split("\n"):
                        max_len = max(max_len, len(line) + 2)
            ws.column_dimensions[col_cells[0].column_letter].width = max(max_len * 1.3, 12)
        ws.freeze_panes = "C2"

    wb.save(output_path)
    print(f"\nExcel 已保存: {output_path}")


def step3_analyze():
    print("=" * 60)
    print(" 第三步: 分析指纹解锁时间")
    print("=" * 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    log_files = glob.glob(os.path.join(FILTERED_DIR, "*fingerprint_filtered.log"))
    if not log_files:
        print(f"在 {FILTERED_DIR} 中未找到过滤后的日志文件")
        return

    all_results = {}
    for lf in sorted(log_files):
        basename = os.path.basename(lf)
        print(f"分析: {basename}")
        unlocks = analyze_log_file(lf)
        valid = [u for u in unlocks if u.get("指纹匹配成功") is not None]
        type_counts = {}
        for u in valid:
            t = u.get("_unlock_type", "未知")
            type_counts[t] = type_counts.get(t, 0) + 1
        type_str = ", ".join(f"{k}={v}" for k, v in type_counts.items())
        print(f"  检测到 {len(unlocks)} 次按压, {len(valid)} 次成功解锁 ({type_str})")
        # 从过滤文件名还原标识: xxx__android_fingerprint_filtered.log -> 压缩包名/原文件名
        short_name = basename.split("__")[0] if "__" in basename else basename[:31]
        all_results[short_name] = valid

    if not any(all_results.values()):
        print("未检测到有效的指纹解锁流程")
        return

    output_path = os.path.join(OUTPUT_DIR, "fingerprint_unlock_analysis.xlsx")
    # 如果文件被占用，自动换文件名
    if os.path.exists(output_path):
        try:
            with open(output_path, "a"):
                pass
        except PermissionError:
            from datetime import datetime as dt
            ts = dt.now().strftime("%H%M%S")
            output_path = os.path.join(OUTPUT_DIR, f"fingerprint_unlock_analysis_{ts}.xlsx")
            print(f"  原文件被占用，保存到: {output_path}")
    write_excel(all_results, output_path)

    # 摘要
    print("\n===== 摘要 =====")
    for sheet, unlocks in all_results.items():
        print(f"\n[{sheet}] 共 {len(unlocks)} 次成功解锁:")
        for i, u in enumerate(unlocks, 1):
            fd = format_ts(u.get("手指按压"))
            base = calc_delta_ms(u.get("手指按压"), u.get("指纹匹配成功"))
            total = calc_delta_ms(u.get("手指按压"), u.get("开始显示应用画面"))
            utype = u.get("_unlock_type", "未知")
            hal = u.get("HAL按压到解锁(ms)", "-")
            kpi = u.get("HAL算法时间(ms)", "-")
            print(f"  第{i}次[{utype}]: {fd} | 底层={base}ms | 总耗时={total}ms | HAL={hal}ms | KPI={kpi}ms")


# ============================================================
#  主入口
# ============================================================

def detect_encoding(filepath: str) -> str:
    """检测文件编码 (BOM 检测)"""
    with open(filepath, "rb") as f:
        head = f.read(4)
    if head[:2] == b"\xff\xfe":
        return "utf-16-le"
    elif head[:2] == b"\xfe\xff":
        return "utf-16-be"
    elif head[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    return "utf-8"


def step2_filter_files(log_files: list[str]):
    """直接过滤指定的 log/txt 文件（跳过解压）"""
    print("=" * 60)
    print(" 过滤指纹相关日志 (直接文件模式)")
    print("=" * 60)

    os.makedirs(FILTERED_DIR, exist_ok=True)

    for log_path in log_files:
        print(f"处理: {log_path}")
        enc = detect_encoding(log_path)
        print(f"  编码: {enc}")
        basename = os.path.splitext(os.path.basename(log_path))[0]
        out_filename = f"{basename}__android_fingerprint_filtered.log"
        dst_path = os.path.join(FILTERED_DIR, out_filename)

        matched = 0
        total = 0
        with open(log_path, "r", encoding=enc, errors="ignore") as fin, \
             open(dst_path, "w", encoding="utf-8") as fout:
            for line in fin:
                total += 1
                if FILTER_PATTERN.search(line):
                    fout.write(line)
                    matched += 1
        print(f"  {total} 行 -> 匹配 {matched} 行 -> {dst_path}")

    print(f"\n过滤完成 -> {FILTERED_DIR}/\n")


def main():
    print("\n🔍 安卓指纹解锁日志一键分析工具\n")

    if len(sys.argv) < 2:
        print("用法:")
        print("  python fingerprint_unlock_speed_analysis.py <file1.zip> [file2.zip ...]     # 解压+分析")
        print("  python fingerprint_unlock_speed_analysis.py <file1.log> [file2.txt ...]    # 直接分析")
        print("  python fingerprint_unlock_speed_analysis.py <file1.zip> <file2.log> ...    # 混合模式")
        sys.exit(1)

    input_files = sys.argv[1:]
    zip_files = []
    log_files = []

    for f in input_files:
        if not os.path.isfile(f):
            print(f"错误: 文件不存在 -> {f}")
            sys.exit(1)
        ext = os.path.splitext(f)[1].lower()
        if ext == ".zip":
            zip_files.append(f)
        elif ext in (".log", ".txt"):
            log_files.append(f)
        else:
            print(f"警告: 不支持的文件类型，跳过 -> {f}")

    if not zip_files and not log_files:
        print("错误: 没有可处理的文件 (支持 .zip / .log / .txt)")
        sys.exit(1)

    print(f"zip 文件: {len(zip_files)} 个, log/txt 文件: {len(log_files)} 个\n")

    # 清空旧的过滤文件，避免混入之前的结果
    if os.path.exists(FILTERED_DIR):
        for f in os.listdir(FILTERED_DIR):
            os.remove(os.path.join(FILTERED_DIR, f))

    # 解压 zip 文件
    if zip_files:
        step1_extract(zip_files)

    # 过滤：从解压目录找 android.log + 直接指定的 log/txt 文件
    if zip_files:
        step2_filter()
    if log_files:
        step2_filter_files(log_files)

    # 分析
    step3_analyze()

    print("\n✅ 全部完成!")
    if zip_files:
        print(f"   解压文件 -> {EXTRACTED_DIR}/")
    print(f"   过滤日志 -> {FILTERED_DIR}/")
    print(f"   分析结果 -> {OUTPUT_DIR}/fingerprint_unlock_analysis.xlsx")


if __name__ == "__main__":
    main()
